import openpyxl


def get_rowcount(filename, sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetname)
    return sheet.max_row


def get_col_count(filename, sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetname)
    return sheet.max_column


def read_data(filename, sheetname, rowno, colno):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetname)
    return sheet.cell(row=rowno, column=colno).value


def write_data(filename, sheetname, rowno, colno, data):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetname)
    sheet.cell(row=rowno, column=colno).value = data
    wb.save(filename)